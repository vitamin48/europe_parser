"""
Скрипт считывает файл JSON (необходимо задать FILE_NAME_JSON и имя результирующего файла) с товарами и записывает
данные в Excel.
Формируется 1 лист с товарами и логикой обработки.
"""
import json
import pandas as pd
from openpyxl.utils import get_column_letter

# Имена файлов для удобства вынесены в константы
FILE_NAME_JSON = 'out/data.json'
RESULT_FILE_NAME = 'out/ПарсЕвро250909.xlsx'


def read_json():
    """Считывает данные из JSON файла."""
    try:
        with open(FILE_NAME_JSON, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
            return data
    except FileNotFoundError:
        print(f"Ошибка: Файл {FILE_NAME_JSON} не найден.")
        return None
    except json.JSONDecodeError:
        print(f"Ошибка: Не удалось декодировать JSON из файла {FILE_NAME_JSON}.")
        return None


def transform_price(x):
    """Трансформирует цену по заданным правилам."""
    if not isinstance(x, (int, float)):
        return 0  # Возвращаем 0 или другое значение по умолчанию, если цена некорректна

    match x:
        case _ if x < 100:
            result = x * 7
        case _ if x < 250:
            result = x * 6
        case _ if x < 500:
            result = x * 5
        case _ if x < 750:
            result = x * 4.5
        case _ if x < 1000:
            result = x * 4
        case _ if x < 1500:
            result = x * 3.5
        case _ if x < 2000:
            result = x * 3
        case _ if x < 3000:
            result = x * 2.5
        case _ if x < 4000:
            result = x * 2
        case _:
            result = x * 1.5

    result = max(result, 590)
    return round(result)


def create_df_by_dict(data_dict):
    """
    Преобразует словарь с данными о товарах в DataFrame, обрабатывая
    описание, вес и другие поля согласно новым требованиям.
    """
    rows = []

    for key, value in data_dict.items():
        characteristics = value.get("characteristics", {})

        # 1. Формирование нового описания
        description_parts = []
        base_description = value.get("description", "")
        if base_description not in ('-', '0', ''):
            description_parts.append(base_description)

        # Добавляем характеристики в описание
        for char_key, char_value in characteristics.items():
            # --- ИЗМЕНЕНИЕ ЗДЕСЬ ---
            # Пропускаем характеристику "Вес", так как она вынесена в отдельный столбец
            if char_key == "Вес":
                continue

            # Добавляем только если значение не является "заглушкой"
            if str(char_value).strip() not in ('-', '0', '--', ''):
                description_parts.append(f"{char_key}: {char_value}")

        full_description = "\n".join(description_parts)

        # 2. Извлечение веса (остается без изменений)
        weight = characteristics.get("Вес", "-")

        # 3. Обработка ссылок на изображения
        img_urls = value.get("img_url", [])
        img_url1 = img_urls[0] if img_urls else "-"
        img_url2 = ", ".join(img_urls[1:]) if len(img_urls) > 1 else "-"

        # Подготовка строки для DataFrame
        row = {
            "ArtNumber": key,
            "Название": value.get("name", ""),
            "Цена Европы": value.get("price", 0),
            "Описание": full_description,
            "Вес товара": weight,
            "Ссылка на главное фото товара": img_url1,
            "Ссылки на другие фото товара": img_url2,
            "art_url": value.get("art_url", ""),
            "Характеристики": str(characteristics)
        }
        rows.append(row)

    df = pd.DataFrame(rows)

    # Добавляем вычисляемые столбцы
    df["Артикул"] = df["ArtNumber"].apply(lambda art: f'e_{art}')
    df['Цена для OZON'] = df['Цена Европы'].apply(transform_price)
    df['Цена до скидки'] = df['Цена для OZON'].apply(lambda x: int(round(x * 1.3)))
    df["НДС"] = "Не облагается"

    # Задаем итоговый порядок столбцов
    desired_order = [
        'Артикул', 'Название', 'Цена для OZON', 'Цена до скидки', 'НДС',
        'Цена Европы', 'Вес товара', 'Ссылка на главное фото товара',
        'Ссылки на другие фото товара', 'Описание', 'ArtNumber',
        'Характеристики', 'art_url'
    ]
    final_columns = [col for col in desired_order if col in df.columns]
    result_df = df[final_columns]

    return result_df


def create_xls(df_res, file_name):
    """Сохраняет итоговый DataFrame в Excel-файл с одним листом и форматированием."""
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        df_res.to_excel(writer, sheet_name='OZON', index=False)
        worksheet_ozon = writer.sheets['OZON']

        # Автоподбор ширины столбцов и форматирование
        for idx, column in enumerate(df_res.columns):
            column_width = max(df_res[column].astype(str).map(len).max(), len(column)) + 2
            col_letter = get_column_letter(idx + 1)
            worksheet_ozon.column_dimensions[col_letter].width = column_width

        # Установка фиксированной ширины для длинных полей
        worksheet_ozon.column_dimensions['B'].width = 30  # Название
        worksheet_ozon.column_dimensions[get_column_letter(df_res.columns.get_loc('Описание') + 1)].width = 50
        worksheet_ozon.column_dimensions[get_column_letter(df_res.columns.get_loc('Характеристики') + 1)].width = 50
        worksheet_ozon.column_dimensions[
            get_column_letter(df_res.columns.get_loc('Ссылка на главное фото товара') + 1)].width = 30
        worksheet_ozon.column_dimensions[
            get_column_letter(df_res.columns.get_loc('Ссылки на другие фото товара') + 1)].width = 30
        worksheet_ozon.column_dimensions[get_column_letter(df_res.columns.get_loc('art_url') + 1)].width = 30

        # Закрепление первой строки
        worksheet_ozon.freeze_panes = 'A2'


if __name__ == '__main__':
    data_json = read_json()
    if data_json:
        df_result = create_df_by_dict(data_dict=data_json)
        create_xls(df_result, file_name=RESULT_FILE_NAME)
        print(f"Файл '{RESULT_FILE_NAME}' успешно создан.")
